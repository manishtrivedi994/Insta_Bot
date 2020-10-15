from selenium.webdriver.common.keys import Keys
import openpyxl as openpyxl
from selenium import  webdriver
import time
import random
import xlsxwriter
from selenium.webdriver.support.ui import Select
import os

driver = webdriver.Chrome(executable_path=r'/home/manish/Downloads/chromedriver')
driver1 = webdriver.Chrome(executable_path=r'/home/manish/Downloads/chromedriver')

#Get Random Email from https://generator.email/
driver1.get("https://www.fakemail.net/")
time.sleep(4)
email = driver1.find_element_by_id('email').text

#Get Random Name from a website
driver.get("https://www.behindthename.com/random/random.php?number=2&sets=1&gender=both&surname=&usage_ind=1")
time.sleep(6)
firstName = driver.find_element_by_xpath('/html/body/div[2]/div/div/center/div[1]/span/a[1]').text
lastName = driver.find_element_by_xpath('/html/body/div[2]/div/div/center/div[1]/span/a[2]').text
#print(firstName+" "+lastName)

driver.get("https://www.instagram.com/accounts/emailsignup/")
time.sleep(5)

#providing email
search = driver.find_element_by_name("emailOrPhone")
search.send_keys(email)

#providing name
search = driver.find_element_by_name("fullName")
search.send_keys(firstName+" "+lastName)

#generating and providing username from name
username = lastName+firstName+(str)(random.randint(0, 999))
search = driver.find_element_by_name("username")
search.send_keys(username)

#providing password
search = driver.find_element_by_name("password")
password = 'instagramBotTestingPassword'
search.send_keys(password)


#click on signup button
button = driver.find_element_by_xpath("/html/body/div[1]/section/main/div/article/div/div[1]/div/form/div[7]/div/button")
button.click()

#providing random date of birth
year = (str)(random.randint(1984, 2002))
month = (str)(random.randint(1, 12))
day = (str)(random.randint(1, 28))
time.sleep(4)
select = Select(driver.find_element_by_xpath('/html/body/div[1]/section/main/div/article/div/div[1]/div/div[4]/div/div/span/span[3]/select'))
select.select_by_value(year)
select = Select(driver.find_element_by_xpath('/html/body/div[1]/section/main/div/article/div/div[1]/div/div[4]/div/div/span/span[2]/select'))
select.select_by_value(day)
select = Select(driver.find_element_by_xpath('/html/body/div[1]/section/main/div/article/div/div[1]/div/div[4]/div/div/span/span[1]/select'))
select.select_by_value(month)

#clicking the next button
button = driver.find_element_by_xpath("/html/body/div[1]/section/main/div/article/div/div[1]/div/div[6]/button")
button.click()

#write to a file
PATH = './instaBots.xlsx'
file = 'instaBots.xlsx'
if os.path.isfile(PATH) and os.access(PATH, os.W_OK):
    new_row = [email, firstName + " " + lastName, username, password]

    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Bots1']  # Older method was  .get_sheet_by_name('Sheet1')
    row = ws.max_row + 1 #getting the maximum row

    for col, entry in enumerate(new_row, start=1):
        ws.cell(row=row, column=col, value=entry)

    wb.save(file)
    wb.close()
else:
    workbook = xlsxwriter.Workbook(file)
    worksheet = workbook.add_worksheet("Bots1")
    worksheet.write('B1', email)
    worksheet.write('A1', firstName + " " + lastName)
    worksheet.write('C1', username)
    worksheet.write('D1', password)
    workbook.close()


#otp retreival
time.sleep(30)
otp = driver1.find_element_by_xpath('/html/body/div[2]/div[3]/div[2]/div[1]/div/table/tbody/tr[3]/td[1]/span[3]').text
otp = otp[:6]
otp_input = driver.find_element_by_xpath('/html/body/div[1]/section/main/div/article/div/div[1]/div[2]/form/div/div[1]/input')
otp_input.send_keys(otp)
button = driver.find_element_by_xpath('/html/body/div[1]/section/main/div/article/div/div[1]/div[2]/form/div/div[2]/button')
button.click()
time.sleep(10)

#closing the popup
driver.find_element_by_xpath('/html/body/div[4]/div/div/div/div[3]/button[2]').click()
time.sleep(4)


#searching for the user to follow
searchBox = driver.find_element_by_xpath('/html/body/div[1]/section/nav/div[2]/div/div/div[2]/input')
searchBox.send_keys("usrname")
time.sleep(4)
searchBox.send_keys(Keys.ENTER)
searchBox.send_keys(Keys.ENTER)
time.sleep(4)
followButton = driver.find_element_by_xpath('/html/body/div[1]/section/main/div/header/section/div[1]/div[1]/div/div/button')
followButton.click()
